import copy
import openpyxl
import numpy as np

# Вывод массива данных
def Print_List(x):
    print('\n')
    for i in range(len(x)):
        print(x[i])

file_path = 'МКП_РГР.xlsx'
workbook = openpyxl.load_workbook(file_path)

# Выбираем лист "Data" или создаем его, если его нет
sheet_name = 'Data'
if sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    sheet.delete_rows(1, sheet.max_row)
else:
    sheet = workbook.create_sheet(title=sheet_name)

text_file_path = 'Данные'

filter_list = [1,4,7,10,13,16,19,22,25,28]
with open(text_file_path, 'r') as file:
    # Читаем строки из файла
    lines = file.readlines()
    # Разбиваем строки на части (предполагаем, что разделены пробелами)
    for line_number, line in enumerate(lines, start=1):
        data_parts = line.split()
        # Записываем данные в соответствующие ячейки в лист "Data"
        for col_number, data_part in enumerate(data_parts, start=1):
            cell = sheet.cell(row=line_number, column=col_number)
            # Пробуем преобразовать данные в числовой формат
            try:
                cell.value = float(data_part.replace(',', '.'))  # Заменяем запятую на точку для правильного преобразования
            except ValueError:
                # Если не удалось преобразовать в число, оставляем строку
                cell.value = data_part

# Запись названия КА в одну ячейку
for i in range(1,31,3):
    k = 1
    name = sheet.cell(row=i,column=k).value
    while sheet.cell(row=i,column=k+1).value != None:
        k += 1
        name = name + ' ' + str(sheet.cell(row=i,column=k).value)
        cell = sheet.cell(row=i, column=k)
        cell.value = None
    cell = sheet.cell(row=i,column=1)
    cell.value = name

# Сохраняем изменения в файле
workbook.save(file_path)
print(f"Данные успешно записаны в лист {sheet_name} файла {file_path}")

# Закрытие книги после использования
workbook.close()

# Создание массива для записи данных КА для подсчетов пересечений

# Запись названий спутников
data_list = []
for i in range(1,29,3):
    name_list = []
    cell = sheet.cell(row = i, column=1)
    name_list.append(cell.value)
    data_list.append(name_list)
del name_list

# Запись данных для каждого спутника
for j in range(len(data_list)):
    for d in range(2,9):
        cell = sheet.cell(row = (j+1)*3, column=d)
        data_list[j].append(cell.value)

# Константы
nu = 396628
R = 6378.16
J2 = 0.0010827
eps = 3/2 * nu * J2 * R**2

# Рассчитываем Омеги, сначала Град/Виток, затем Град/Сутки
for i in range(len(data_list)):
    omega_1 = -(2*np.pi*eps) / nu / (((nu * (1 / (data_list[i][7] / 86400))**2/4/np.pi**2)**(1/3)) * (1-data_list[i][4]**2))**2 * np.cos(data_list[i][2]*np.pi/180)
    data_list[i].append(omega_1)
    omega_2 = data_list[i][-1]/(1 / (data_list[i][7] / 86400)) * 86400
    data_list[i].append(omega_2)

# Поиск КЛА с наименьшей высотой
# Для этого ищем наибольшее значение в data_list[n][7]
max = 0
min_ID = 0
for i in range(len(data_list)):
    if max<data_list[i][7]:
        max = data_list[i][7]
        min_ID = i

# Рассчитываем и записываем начальные и конечные координаты спутников
coordinate_list = []
for i in range(len(data_list)):
    x_0 = data_list[i][3]
    # 1826 это количество дней за 5 лет
    x_k = x_0 + (data_list[i][-1]-data_list[min_ID][-1]) * 1826
    info_list = [data_list[i][0],x_0,x_k]
    coordinate_list.append(info_list)

def find_intersection(KA1, KA2, i, j):
    # Находим наклон и точку пересечения для первой линии
    y1 = KA1[1]
    y2 = KA1[2]
    y3 = KA2[1]
    y4 = KA2[2]
    x1 = 1826
    x0 = 0
    m1 = (y2 - y1) / (x1 - x0)
    b1 = y1 - m1 * x0
    # Находим наклон и точку пересечения для второй линии
    m2 = (y4 - y3) / (x1 - x0)
    b2 = y3 - m2 * x0
    # Решаем систему уравнений для нахождения x и y точки пересечения
    x_intersect = (b2 - b1) / (m1 - m2)
    y_intersect = m1 * x_intersect + b1

    if x_intersect < 0 or x_intersect > 1826:
        return None
    else:
        return x_intersect, y_intersect, KA1[0], KA2[0], i, j

cross_list = []
for i in range(len(coordinate_list)):
    for j in range(i+1,len(coordinate_list)):
        cross_list.append(find_intersection(coordinate_list[i],coordinate_list[j], i, j))

cross_c = copy.deepcopy(cross_list)
for i in range(len(cross_list)):
    if cross_list[i] == None:
        cross_c.remove(None)
cross_list = cross_c
del cross_c
# Заполнение таблицы эксель данными КА которые пересекаются

# Выбираем лист "Data" или создаем его, если его нет
workbook = openpyxl.load_workbook(file_path)

# Выбираем лист "Crossing" или создаем его, если его нет
sheet_name = 'Crossing'
if sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    sheet.delete_rows(1, sheet.max_row)
else:
    sheet = workbook.create_sheet(title=sheet_name)

name_list = ["Пересечения","Δa","Δω","Δv1","Δv2","Δv Σдек"]
for i in range(len(name_list)):
    cell = sheet.cell(row=1, column=i+1)
    cell.value = name_list[i]

k = 2
for i in range(len(cross_list)):

        cell = sheet.cell(row=k,column=1)
        cell.value = cross_list[i][2] + " –> " + cross_list[i][3]
        a1 = ((nu * (1 / (data_list[cross_list[i][4]][7] / 86400)) ** 2 / 4 / np.pi ** 2) ** (1 / 3))
        a2 = ((nu * (1 / (data_list[cross_list[i][5]][7] / 86400)) ** 2 / 4 / np.pi ** 2) ** (1 / 3))
        # Запись дельта а
        cell = sheet.cell(row=k,column=2)
        cell.value = abs(a1 - a2)
        # Запись дельта w
        cell = sheet.cell(row=k,column=3)
        cell.value = abs(data_list[cross_list[i][4]][5] - data_list[cross_list[i][5]][5])
        # Запись дельта v1
        cell = sheet.cell(row=k,column=4)
        cell.value = (2*nu*a2/a1/(a1+a2))**(1/2)-(nu/a1)**(1/2)
        # Запись дельта v2
        cell = sheet.cell(row=k,column=5)
        cell.value = (nu/a2)**(1/2)-(2*nu*a1/a2/(a1+a2))**(1/2)
        # Запись Δv Σдек
        cell = sheet.cell(row=k,column=6)
        cell.value = abs(sheet.cell(row=k,column=5).value + sheet.cell(row=k,column=4).value)
        # # Запись Δv Σпол
        # cell = sheet.cell(row=k, column=7)
        # cell.value = (sheet.cell(row=k,column=4).value**2+sheet.cell(row=k,column=5).value**2)**(1/2)
        k += 1

        # Делаем тоже самое но от второго КА к первому КА
        cell = sheet.cell(row=k,column=1)
        cell.value = cross_list[i][3] + " –> " + cross_list[i][2]
        a1 = ((nu * (1 / (data_list[cross_list[i][5]][7] / 86400)) ** 2 / 4 / np.pi ** 2) ** (1 / 3))
        a2 = ((nu * (1 / (data_list[cross_list[i][4]][7] / 86400)) ** 2 / 4 / np.pi ** 2) ** (1 / 3))
        # Запись дельта а
        cell = sheet.cell(row=k, column=2)
        cell.value = abs(a1 - a2)
        # Запись дельта w
        cell = sheet.cell(row=k, column=3)
        cell.value = abs(data_list[cross_list[i][5]][5] - data_list[cross_list[i][4]][5])
        # Запись дельта v1
        cell = sheet.cell(row=k, column=4)
        cell.value = (2 * nu * a2 / a1 / (a1 + a2)) ** (1 / 2) - (nu / a1) ** (1 / 2)
        # Запись дельта v2
        cell = sheet.cell(row=k, column=5)
        cell.value = (nu / a2) ** (1 / 2) - (2 * nu * a1 / a2 / (a1 + a2)) ** (1 / 2)
        # Запись Δv Σдек
        cell = sheet.cell(row=k, column=6)
        cell.value = abs(sheet.cell(row=k, column=5).value + sheet.cell(row=k, column=4).value)
        # # Запись Δv Σпол
        # cell = sheet.cell(row=k, column=7)
        # cell.value = (sheet.cell(row=k, column=4).value ** 2 + sheet.cell(row=k, column=5).value ** 2) ** (1 / 2)
        k += 1

# Сохраняем изменения в файле
workbook.save(file_path)

# Вывод массива данных КА
Print_List(data_list)
Print_List(coordinate_list)
Print_List(cross_list)
# Закрытие книги после использования
workbook.close()

